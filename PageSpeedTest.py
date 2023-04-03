import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from openpyxl.styles import Font, PatternFill
import datetime

now = datetime.datetime.now()
date_string = now.strftime("%Y-%m-%d_%H-%M-%S")

# Slack Bot token and channel ID
slack_token = "YOUR_SLACK_BOT_API"
channel_id = "YOUR_SLACK_CHANNEL_ID"

# Authenticate the Slack client
client = WebClient(token=slack_token)


# Set up authentication
credentials = service_account.Credentials.from_service_account_file(
    'C:\\Users\\Anonymous\\Downloads\\ultra-depot-367306-e60d1a76d5d8.json')
service = build('pagespeedonline', 'v5', credentials=credentials)

urls = [

    'https://example.com',
    'https://example.com/page1'
]

# Define parameters for PageSpeed Insights API request
metrics_desktop = {
    'first-contentful-paint': 3,
    'largest-contentful-paint': 4,
    'cumulative-layout-shift': 0.25,
    'speed-index': 5.8,
    'total-blocking-time': 600
}

metrics_mobile = {
    'first-contentful-paint': 3,
    'largest-contentful-paint': 4,
    'cumulative-layout-shift': 0.25,
    'speed-index': 5.8,
    'total-blocking-time': 600
}

# Set parameters for PageSpeed Insights API request
params_desktop = {
    'url': '',
    'strategy': 'DESKTOP',
    'category': ["PERFORMANCE", "ACCESSIBILITY", "SEO", "BEST_PRACTICES"],
    'locale': 'en_US'
}

params_mobile = {
    'url': '',
    'strategy': 'MOBILE',
    'category': ["PERFORMANCE", "ACCESSIBILITY", "SEO", "BEST_PRACTICES"],
    'locale': 'en_US'
}

# Create or open an existing Excel file
red_font = Font(color='FF0000')
red_fill = PatternFill(fill_type='solid', start_color='FFC7CE', end_color='FFC7CE')

filename = f"Results_{date_string}.xlsx"
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet["A1"] = "URL"
for i, metric_name in enumerate(metrics_desktop.keys()):
    sheet.cell(row=1, column=i+2, value=f"Desktop {metric_name}")
for i, metric_name in enumerate(metrics_mobile.keys()):
    sheet.cell(row=1, column=i+len(metrics_desktop)+2, value=f"Mobile {metric_name}")

# Run PageSpeed Insights API request and add results to the Excel file
for row, url in enumerate(urls, start=2):
    # Update URL in test parameters
    params_desktop['url'] = url
    params_mobile['url'] = url

    result_desktop = service.pagespeedapi().runpagespeed(**params_desktop).execute()
    result_mobile = service.pagespeedapi().runpagespeed(**params_mobile).execute()

    # Extract metric values from result dictionary
    desktop_results = []
    mobile_results = []
    if result_desktop is not None:
        lighthouse_result = result_desktop['lighthouseResult']
        for metric_name, threshold in metrics_desktop.items():
            try:
                metric_value_str = lighthouse_result['audits'][metric_name]['displayValue']
                metric_value = float(metric_value_str.replace(',', '').replace('\xa0s', '').replace('\xa0ms', ''))
                desktop_results.append(metric_value)
            except KeyError:
                desktop_results.append(None)
    if result_mobile is not None:
        lighthouse_result = result_mobile['lighthouseResult']
        for metric_name, threshold in metrics_mobile.items():
            try:
                metric_value_str = lighthouse_result['audits'][metric_name]['displayValue']
                metric_value = float(metric_value_str.replace(',', '').replace('\xa0s', '').replace('\xa0ms', ''))
                mobile_results.append(metric_value)
            except KeyError:
                mobile_results.append(None)
    # Add URL and metric values to the Excel file
    sheet.cell(row=row, column=1, value=url)
    for i, metric_value in enumerate(desktop_results):
        sheet.cell(row=row, column=i + 2, value=metric_value)
        if metric_value is not None and metric_value > metrics_desktop[list(metrics_desktop.keys())[i]]:
            sheet.cell(row=row, column=i + 2).font = red_font
            sheet.cell(row=row, column=i + 2).fill = red_fill
    for i, metric_value in enumerate(mobile_results):
        sheet.cell(row=row, column=i + len(metrics_desktop) + 2, value=metric_value)
        if metric_value is not None and metric_value > metrics_mobile[list(metrics_mobile.keys())[i]]:
            sheet.cell(row=row, column=i + len(metrics_desktop) + 2).font = red_font
            sheet.cell(row=row, column=i + len(metrics_desktop) + 2).fill = red_fill

# Save the Excel file
workbook.save(filename)

# Send the results to the Slack channel
try:
    client.files_upload_v2(
        channel=channel_id,
        file=filename,
        title="PageSpeed Insights Results",
        headers={"Authorization": f"Bearer {slack_token}"}
    )
except SlackApiError as e:
    print("Error uploading file: {}".format(e))